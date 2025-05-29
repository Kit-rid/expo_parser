import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
import os

def encrypt_excel_file(input_file):
    """
    Encrypt Excel file content after first 5 company rows (9 rows total including header)
    Keep 'Название' and 'Рубрика' columns unencrypted
    """
    print(f"Обработка файла: {os.path.basename(input_file)}")
    
    # Load the workbook
    wb = load_workbook(input_file)
    ws = wb.active
    
    # Get the last row with data
    last_row = ws.max_row
    
    # Get column indices for 'Название' and 'Рубрика' (they are in columns A and B)
    unencrypted_columns = [1, 2]  # A and B columns
    
    # Starting from row 10 (after 4 header rows + 5 company rows)
    for row in range(10, last_row + 1):
        # Process all columns
        for col in range(1, ws.max_column + 1):
            # Skip 'Название' and 'Рубрика' columns
            if col not in unencrypted_columns:
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    # Replace any value with exactly 4 X's
                    cell.value = 'XXXX'
                    
                    # Preserve the font and alignment
                    cell.font = Font(name='Times New Roman')
                    cell.alignment = Alignment(horizontal='left')
    
    # Create 'демо_версии' directory if it doesn't exist
    os.makedirs('демо_версии', exist_ok=True)
    
    # Generate output filename
    file_name = os.path.basename(input_file)
    output_file = os.path.join('демо_версии', f'демо_версия_{file_name}')
    
    # Save the modified workbook
    wb.save(output_file)
    print(f"Создан файл: {output_file}")

def main():
    # Get list of Excel files in the excel directory
    excel_dir = 'excel'
    if not os.path.exists(excel_dir):
        print("Папка 'excel' не найдена!")
        return
        
    excel_files = [f for f in os.listdir(excel_dir) if f.endswith('.xlsx')]
    
    if not excel_files:
        print("Excel файлы не найдены в папке 'excel'!")
        return
    
    print(f"Найдено файлов для обработки: {len(excel_files)}")
    print("Колонки 'Название' и 'Рубрика' останутся без изменений, остальные данные будут зашифрованы")
    
    # Process all files
    for file in excel_files:
        try:
            input_file = os.path.join(excel_dir, file)
            encrypt_excel_file(input_file)
        except Exception as e:
            print(f"Ошибка при обработке файла {file}: {e}")
    
    print("\nОбработка завершена!")
    print(f"Обработано файлов: {len(excel_files)}")
    print("Зашифрованные файлы сохранены в папку 'демо_версии'")

if __name__ == '__main__':
    main() 