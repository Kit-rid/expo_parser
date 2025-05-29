import requests
from bs4 import BeautifulSoup
import json
from urllib.parse import urljoin
import time
import pandas as pd
import os
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image

def create_top_rows(worksheet):
    # Создаем первые 4 строки с контентом
    rows = [
        ["базы данных предприятий России", "www.базы-предприятий.рф", "SMS рассылка: 2,99 р. / сообщение", "", 
         "Здесь может быть Ваша реклама", "", "", "Обновляется февраль и сентябрь", "", ""],
        ["", "helper@aitera.com", "E-mail адресная рассылка: 3 р. 99 коп. / письмо", "", "", "", "", "", "", ""],
        ["", "\nтелефон: 8495 223 35 57", "Холодные звонки: 17 990 р. / 500 диалогов", "", "", "", "", "", "", ""],
        ["Название", "Рубрика", 
         "Телефоны", "Email", "Сайт"]
    ]
    logo = Image('Рисунок1.jpg')
    logo.width = 150
    logo.height = 50

    # Создаем стиль для заголовков
    header_fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
    times_new_roman_bold = Font(name='Times New Roman', bold=True)
    

    # Устанавливаем высоту для первых 4 строк (30 пикселей ≈ 22.5 пунктов)
    for row in range(1, 5):
        worksheet.row_dimensions[row].height = 22.5

    # Заполняем строки
    for row_idx, row_data in enumerate(rows, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name='Calibri', bold=True)
            
            # Форматирование заголовков
            if row_idx == 4:
                cell.font = times_new_roman_bold  # Жирный Times New Roman для заголовков
                cell.alignment = Alignment(horizontal='center')
                cell.fill = header_fill
            
            # Форматирование специальных ячеек
            if row_idx == 1:
                if col_idx == 3:  # SMS рассылка
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif col_idx == 5:  # Здесь может быть Ваша реклама
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='Calibri', bold=True, color='1F497D', size=16)
    
    # Объединение ячеек для первой строки
    worksheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    worksheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
    worksheet.merge_cells(start_row=1, start_column=5, end_row=2, end_column=6)
    worksheet.add_image(logo, 'A1')
    
    # Закрепляем первые 4 строки
    worksheet.freeze_panes = "A5"

def get_company_details(url, headers):
    """Получение детальной информации о компании из dl-horizontal"""
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'lxml')
        dl_horizontal = soup.find(class_='dl-horizontal')
        
        if not dl_horizontal:
            print(f"Не найден dl-horizontal на странице {url}")
            return {}
            
        company_info = {}
        
        # Находим нужные поля по их названиям
        dt_elements = dl_horizontal.find_all('dt')
        dd_elements = dl_horizontal.find_all('dd')
        
        for dt, dd in zip(dt_elements, dd_elements):
            key = dt.get_text(strip=True)
            
            # Сопоставляем поля с нужными нам данными
            if key == "Сайт:":
                site_link = dd.find('a')
                company_info['Сайт'] = site_link.get_text(strip=True) if site_link else ''
            elif key == "Телефон:":
                company_info['Телефон'] = dd.get_text(strip=True)
            elif key == "E-mail:":
                email_link = dd.find('a')
                company_info['E-mail'] = email_link.get_text(strip=True) if email_link else ''
            
            # Находим все span элементы с классом label label-primary
            spans = dd.find_all('span', class_='label label-primary')
            # Извлекаем только текст из каждого span
            categories = [span.get_text(strip=True) for span in spans]
            company_info['Рубрика'] = '; '.join(categories)
                
        return company_info
        
    except requests.RequestException as e:
        print(f"Ошибка при получении данных компании с {url}: {e}")
        return {}

def get_table_links(url, headers):
    """Получение ссылок на компании из таблицы на странице /list"""
    try:
        # Добавляем /list к URL, если его там нет
        if not url.endswith('/list'):
            url = f"{url}/list"
            
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'lxml')
        
        # Сначала ищем таблицу fresh-table
        fresh_table = soup.find(id='fresh-table')
        if not fresh_table:
            print(f"Таблица fresh-table не найдена на {url}")
            return []
            
        # Находим все элементы tr в таблице
        table_rows = fresh_table.find_all('tr')
        
        results = []
        
        for row in table_rows:
            # Получаем первый элемент td в строке
            first_td = row.find('td')
            if first_td:
                # Ищем ссылку в td
                link = first_td.find('a')
                if link:
                    href = link.get('href', '')
                    text = link.get_text(strip=True)
                    full_url = urljoin(url, href) if href else ''
                    if full_url:
                        # Получаем детали компании сразу
                        print(f"Получение данных компании: {text}")
                        company_details = get_company_details(full_url, headers)
                        
                        company_data = {
                            'text': text,
                            'url': full_url,
                            'details': company_details
                        }
                        results.append(company_data)
                        time.sleep(0.5)  # Небольшая задержка между запросами
        
        return results
    except requests.RequestException as e:
        print(f"Ошибка при получении ссылок из таблицы {url}: {e}")
        return []

def get_exhibition_links(url, headers):
    """Получение ссылок на выставки с главной страницы"""
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.text, 'lxml')
        links = soup.find_all(class_='list-group-item list-group-item-action')
        
        results = []
        
        for link in links:
            href = link.get('href', '')
            text = link.get_text(strip=True)
            
            # Проверяем, содержит ли текст 2025 и 2024
            if not any(year in text for year in ['2024', '2025']):
                continue
                
            full_url = urljoin(url, href) if href else ''
            if full_url:
                results.append({
                    'text': text,
                    'url': full_url
                })
        
        return results
    except requests.RequestException as e:
        print(f"Ошибка при получении ссылок на выставки: {e}")
        return []

def save_to_excel(exhibition_name, companies_data):
    """Save companies data to Excel file with custom formatting"""
    # Create DataFrame
    excel_data = []
    
    for company in companies_data:
        details = company.get('details', {})
        excel_data.append({
            'Название': company['text'],
            'Рубрика': details.get('Рубрика', ''),
            'Телефоны': details.get('Телефон', ''),
            'Email': details.get('E-mail', ''),
            'Сайт': details.get('Сайт', '')
        })
    
    df = pd.DataFrame(excel_data)
    
    # Create 'excel' directory if it doesn't exist
    os.makedirs('excel', exist_ok=True)
    filename = f'excel/Аитэра +7495 223 35 57 участники выставки {exhibition_name}.xlsx'
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # Сначала записываем пустой DataFrame чтобы создать лист
        pd.DataFrame().to_excel(writer, sheet_name='Аитэра 7495 223 35 57')
        
        # Получаем объект листа
        worksheet = writer.sheets['Аитэра 7495 223 35 57']
        
        # Создаем верхние строки
        create_top_rows(worksheet)
        
        # Записываем данные компаний начиная с 5-й строки
        times_new_roman = Font(name='Times New Roman')
        for row_idx, row in enumerate(df.itertuples(), 5):
            for col_idx, value in enumerate(row[1:], 1):
                cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                cell.font = times_new_roman  # Устанавливаем Times New Roman для всех данных
        
        # Настраиваем ширину колонок
        column_widths = {
            'A': 35, 'B': 25, 'C': 20, 
            'D': 20, 'E': 10, 'F': 35,
            'G': 35, 'H': 25, 'I': 25, 'J': 30
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
            
        # Добавляем возможность сортировки (AutoFilter)
        last_row = len(excel_data) + 4  # 4 - это количество строк в шапке
        worksheet.auto_filter.ref = f"A4:E{last_row}"
    
    print(f"Excel file saved: {filename}")

def parse_expocentr():
    # URL сайта
    url = 'https://icatalog.expocentr.ru/ru'
    
    # Заголовки для имитации браузера
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    
    try:
        print("Шаг 1: Получение ссылок на выставки...")
        exhibition_links = get_exhibition_links(url, headers)
        
        if not exhibition_links:
            print("Ссылки на выставки не найдены!")
            return
            
        print(f"Найдено {len(exhibition_links)} выставок на 2024 и 2025 год")
        
        # Создаем структуру для итогового результата
        result = []
        
        # Обрабатываем каждую ссылку на выставку
        for exhibition in exhibition_links:
            exhibition_data = {
                'exhibition_name': exhibition['text'],
                'exhibition_url': exhibition['url'],
                'companies': []
            }
            
            # Получаем ссылки на компании со страницы выставки
            print(f"\nПолучение компаний с выставки: {exhibition['text']}")
            company_links = get_table_links(exhibition['url'], headers)
            exhibition_data['companies'] = company_links
            result.append(exhibition_data)
            
            # Сохраняем в Excel
            save_to_excel(exhibition['text'], company_links)
            
            print(f"Найдено {len(company_links)} компаний")
            
            # Небольшая задержка между запросами к серверу
            time.sleep(1)
        
        # Сохраняем в JSON файл
        with open('expo_links.json', 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
            
        print(f"\nУспешно обработано {len(result)} выставок")
        total_companies = sum(len(expo['companies']) for expo in result)
        print(f"Всего собрано компаний: {total_companies}")
        
    except Exception as e:
        print(f"Произошла ошибка: {e}")

if __name__ == '__main__':
    parse_expocentr() 