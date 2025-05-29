import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def analyze_excel_file(file_path):
    """Analyze a single Excel file and return statistics."""
    # Read the Excel file starting from the 5th row where the actual data begins
    df = pd.read_excel(file_path, header=3)  # header=3 means use row 4 (0-based) as headers
    
    # Print column names for debugging
    print(f"Columns in file {os.path.basename(file_path)}:")
    print(df.columns.tolist())
    
    # Get exhibition name from filename
    exhibition_name = os.path.basename(file_path).replace('участники выставки ', '').replace('.xlsx', '')
    
    # Count companies with different contact information
    # Using exact column names from the original Excel file
    companies_with_phones = df[df['Телефоны'].notna() & (df['Телефоны'] != '')].shape[0]
    companies_with_websites = df[df['Сайт'].notna() & (df['Сайт'] != '')].shape[0]
    companies_with_emails = df[df['Email'].notna() & (df['Email'] != '')].shape[0]
    total_companies = len(df)
    
    return {
        'Выставка': exhibition_name,
        'Всего компаний': total_companies,
        'Компании с телефонами': companies_with_phones,
        'Компании с сайтами': companies_with_websites,
        'Компании с email': companies_with_emails,
        'Процент с телефонами': f"{(companies_with_phones/total_companies*100):.1f}%" if total_companies > 0 else "0%",
        'Процент с сайтами': f"{(companies_with_websites/total_companies*100):.1f}%" if total_companies > 0 else "0%",
        'Процент с email': f"{(companies_with_emails/total_companies*100):.1f}%" if total_companies > 0 else "0%"
    }

def create_summary_excel(results):
    """Create a summary Excel file with analysis results."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Аитэра 7495 223 35 57"
    
    # Define headers
    headers = [
        'Выставка',
        'Всего компаний',
        'Компании с телефонами',
        'Процент с телефонами',
        'Компании с сайтами',
        'Процент с сайтами',
        'Компании с email',
        'Процент с email'
    ]
    
    # Style for headers
    header_font = Font(name='Times New Roman', bold=True)
    header_fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1', fill_type='solid')
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
    
    # Write data
    times_new_roman = Font(name='Times New Roman')
    for row, result in enumerate(results, 2):
        ws.cell(row=row, column=1, value=result['Выставка']).font = times_new_roman
        ws.cell(row=row, column=2, value=result['Всего компаний']).font = times_new_roman
        ws.cell(row=row, column=3, value=result['Компании с телефонами']).font = times_new_roman
        ws.cell(row=row, column=4, value=result['Процент с телефонами']).font = times_new_roman
        ws.cell(row=row, column=5, value=result['Компании с сайтами']).font = times_new_roman
        ws.cell(row=row, column=6, value=result['Процент с сайтами']).font = times_new_roman
        ws.cell(row=row, column=7, value=result['Компании с email']).font = times_new_roman
        ws.cell(row=row, column=8, value=result['Процент с email']).font = times_new_roman
    
    # Adjust column widths
    column_widths = {
        'A': 40,  # Выставка
        'B': 15,  # Всего компаний
        'C': 20,  # Компании с телефонами
        'D': 20,  # Процент с телефонами
        'E': 20,  # Компании с сайтами
        'F': 20,  # Процент с сайтами
        'G': 20,  # Компании с email
        'H': 20,  # Процент с email
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Add autofilter
    ws.auto_filter.ref = f"A1:H{len(results) + 1}"
    
    # Save the workbook
    wb.save('excel/анализ_выставок.xlsx')
    print("Анализ сохранен в файл: excel/анализ_выставок.xlsx")

def main():
    # Create excel directory if it doesn't exist
    os.makedirs('excel', exist_ok=True)
    
    # Get all Excel files in the excel directory
    excel_files = [f for f in os.listdir('excel') if f.startswith('участники выставки') and f.endswith('.xlsx')]
    
    if not excel_files:
        print("Не найдены файлы Excel для анализа в папке 'excel'")
        return
    
    # Analyze each file
    results = []
    for file in excel_files:
        print(f"Анализируем файл: {file}")
        file_path = os.path.join('excel', file)
        try:
            result = analyze_excel_file(file_path)
            results.append(result)
        except Exception as e:
            print(f"Ошибка при анализе файла {file}: {e}")
    
    if results:
        create_summary_excel(results)
    else:
        print("Нет данных для анализа")

if __name__ == '__main__':
    main() 